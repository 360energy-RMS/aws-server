# Writes data to XLSX file
from openpyxl import load_workbook

class DATA_WRITER:
    
    def writeToFile(data):
        # writes the given data to xlsx file
        filePath = r"C:/Users/RMS-SERVER/Documents/RMS-LOG/RMS_LOG.xlsx"
        workbook = load_workbook(filePath)

        columnDict = {
            0 : 'A',
            1 : 'B',
            2 : 'C',
            3 : 'D',
            4 : 'E',
        }

        sheet = workbook.active
        emptyRow = DATA_WRITER.findEmpty(1)
        i = -1
       
        while i < len(data) - 1:
            i += 1
            sheet[columnDict[i] + emptyRow] = data[i]
        
        workbook.save(filePath)


    def findEmpty(columnToCheck):
        # used by writeToFile to find a blank cell to write data in
        count = 1
        filePath = r"C:/Users/RMS-SERVER/Documents/RMS-LOG/RMS_LOG.xlsx"
        workbook = load_workbook(filePath)
        sheet = workbook.active

        while True:
            count += 1
            cellObj = sheet.cell(row=count, column=columnToCheck)

            if cellObj.value is None:

                return str(count)